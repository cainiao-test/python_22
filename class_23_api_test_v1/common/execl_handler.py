"""
第一：excel类封装需要提供以下功能：

1、选择表单功能
2,读取一个单元格的数据功能
3、读取一行数据  功能
4、读取表单中所有数据功能
5、往单元格中写入数据功能
6、保存数据功能
"""
from openpyxl import load_workbook

# 全局，常量
FILE_NAME = 'data.xlsx'


class ExcelHandler:
    """excel 封装"""

    # 项目来说可能不变
    # 测试数据，写代码之前之前已经写好测试数据的Excel. 文件名
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.sheet = self.wb.worksheets[self.sheet_name]
        else:
            self.sheet = self.wb[self.sheet_name]
    # def choose_sheet(self, sheet_name):
    #     """选择表单.
    #     sheet_name 是整数，根据索引获取。
    #     如果是字符串，根据名字获取 '20190920'
    #     """
    #     if isinstance(sheet_name, int):
    #         return self.wb.worksheets[sheet_name]
    #     return self.wb[sheet_name]

    def headers(self):
        """获取标题"""
        self.open()
        headers = [c.value for c in self.sheet[1]]
        self.wb.close()
        return headers

    def read(self, start_row=2, start_column=1):
        """获取所有的数据"""
        self.open()
        sheet = self.sheet

        header = [c.value for c in sheet[1]]
        self.wb.close()
        # max_row, max column
        data = []
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        return data

    def save(self):
        """保存"""
        self.wb.save(self.file_name)
        self.wb.close()

    def write(self, row, column, data):
        self.open()
        wb = load_workbook(self.file_name)
        self.sheet.cell(row, column).value = data
        # 保存关闭
        wb.save(self.file_name)
        wb.close()


class ExcelHandler2:
    """excel 封装"""

    # 项目来说可能不变
    # 测试数据，写代码之前之前已经写好测试数据的Excel. 文件名
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)
        # if isinstance(sheet_name, int):
        #     self.sheet =  self.wb.worksheets[sheet_name]
        # else:
        #     self.sheet = self.wb.get_sheet_by_name(sheet_name)

    def choose_sheet(self, sheet_name):
        """选择表单.
        sheet_name 是整数，根据索引获取。
        如果是字符串，根据名字获取 '20190920'
        """
        if isinstance(sheet_name, int):
            return self.wb.worksheets[sheet_name]
        return self.wb[sheet_name]

        # 索引

    def read_line(self, sheet_name, line):
        """获取行"""
        sheet = self.choose_sheet(sheet_name)
        sheet_data = sheet[line]
        # 元组 （Cell(1,1), Cell(1,2）
        data = []
        for c in sheet_data:
            data.append(c.value)
        return data

    def read(self, sheet_name, start_row=2, start_column=1):
        """获取所有的数据"""
        sheet = self.choose_sheet(sheet_name)
        # max_row, max column
        data = []
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            data.append(row_data)
        return data

    def read_cell(self, sheet_name, row, column):
        """一个单元格的数据"""
        sheet = self.choose_sheet(sheet_name)
        return sheet.cell(row, column).value

    def save(self):
        """保存"""
        self.wb.save(self.file_name)
        self.wb.close()

    @staticmethod
    def write(file_name, sheet_name, row, column, data):
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        sheet.cell(row, column).value = data
        # 保存关闭
        wb.save(file_name)
        wb.close()


if __name__ == '__main__':
    xls = ExcelHandler(r'/class_23_api_test_v1/data/cases.xlsx')
    print(xls.read())